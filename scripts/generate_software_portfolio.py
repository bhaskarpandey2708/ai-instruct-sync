#!/usr/bin/env python3
"""
Generate Software Opportunity Research Portfolio (Excel + CSV + INDEX).
Working dir: /Users/bhaskar_pandey/Documents/development
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path("/Users/bhaskar_pandey/Documents/development")
XLSX_PATH = ROOT / "Software_Opportunity_Research_Portfolio.xlsx"
CSV_PATH = ROOT / "Software_Opportunity_Master_Portfolio.csv"
INDEX_PATH = ROOT / "SOFTWARE_PORTFOLIO_INDEX.md"

TODAY = date(2026, 7, 11).isoformat()

# ---------------------------------------------------------------------------
# Styles (financial modeling conventions)
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
BLUE_INPUT = Font(name=FONT_NAME, size=10, color="0000FF")  # inputs
BLACK = Font(name=FONT_NAME, size=10, color="000000")
BLACK_BOLD = Font(name=FONT_NAME, size=10, bold=True, color="000000")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, bold=True, color="1F4E79")
SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
LINK_BLUE = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")
FORMULA_FONT = Font(name=FONT_NAME, size=10, color="000000")  # calculated

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_YELLOW = PatternFill("solid", fgColor="FFFF99")  # editable inputs
FILL_LIGHT = PatternFill("solid", fgColor="D6EAF8")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")
FILL_NOW = PatternFill("solid", fgColor="C6EFCE")
FILL_NEXT = PatternFill("solid", fgColor="FFEB9C")
FILL_LATER = PatternFill("solid", fgColor="BDD7EE")
FILL_PARK = PatternFill("solid", fgColor="D9D9D9")
FILL_README = PatternFill("solid", fgColor="E2EFDA")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN


def apply_arial(ws, min_row=1, max_row=200, min_col=1, max_col=30) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.font is None or cell.font.name != FONT_NAME:
                # preserve color/bold if already set via explicit assignment later
                pass
            if cell.value is not None and (cell.font is None or not cell.font.name):
                cell.font = BLACK


# ---------------------------------------------------------------------------
# Opportunity data (28 items P01–P28)
# Scores: Burn, Software-native, Ship speed, Distribution, Moat (1–5)
# ---------------------------------------------------------------------------
# Columns for master (logical):
# id, name, category, one_liner, problem, burn, soft_native, ship, dist, moat,
# status, stage, platforms, repo, tags, notes

OPPS = [
    {
        "id": "P01",
        "name": "instruct-sync",
        "category": "AI Tooling",
        "one_liner": "Sync AI coding rules/instructions across Cursor, Copilot, Claude, Windsurf, Aider, Gemini, Continue",
        "problem": "Every AI coding agent uses different rules files; they drift and teams lose consistency.",
        "burn": 5,
        "soft_native": 5,
        "ship": 5,
        "dist": 5,
        "moat": 3,
        "status": "Public beta",
        "stage": "Shipped / iterate",
        "platforms": "CLI",
        "repo": "instruct-sync",
        "tags": "AI,CLI,devtools,rules,sync,suite",
        "notes": "npm ai-instruct-sync@0.2.0-beta.0; hygiene suite #1 (sync); primary NOW product",
        # deep research
        "market": "AI coding assistants exploding; multi-agent workplaces common. Rules/context fragmentation is daily pain.",
        "competitors": "rulesync, personal scripts, agent-specific converters, Cursor project rules only",
        "why_now": "Agent sprawl (Cursor+Copilot+Claude Code etc.) peaking; OSS CLI distribution via npx is proven.",
        "risks": "Agents change file formats; big vendors ship native sync; commodity CLI risk",
        "evidence": "Live beta, multi-agent adapters, dry-run/backup UX; HN/Reddit-ready narrative",
        "next_research": "Download retention, paid team features demand, enterprise rules governance",
        # build
        "mvp": "status/diff/sync/validate/init across 7+ agents; dry-run default; --json",
        "stack": "TypeScript, Node 20+, zero runtime deps, vitest",
        "effort_weeks": 2,
        "delivery_notes": "Already shipping; focus on agents, lint, team mode",
        # monetization
        "icp": "Indie/pro AI-assisted developers; later eng leads at multi-agent shops",
        "pricing": "OSS free; optional Pro team sync + policy ($8–15/dev/mo later)",
        "gtm": "npx, Show HN, r/cursor, X, npm, directories, compare posts",
        "revenue_model": "Open core → team SaaS",
        "arpu_hint": "$0 OSS; $10–20/seat Pro (est.)",
    },
    {
        "id": "P02",
        "name": "ai-setup-doctor",
        "category": "AI Tooling",
        "one_liner": "Read-only diagnostics for agents, MCP, secrets, Node/project hygiene before AI coding sessions",
        "problem": "Broken MCP, leaked secrets in agent configs, bloated rules, and project hygiene waste hours.",
        "burn": 5,
        "soft_native": 5,
        "ship": 5,
        "dist": 4,
        "moat": 3,
        "status": "Published beta",
        "stage": "Shipped / iterate",
        "platforms": "CLI",
        "repo": "ai-setup-doctor",
        "tags": "AI,CLI,diagnostics,MCP,secrets,suite",
        "notes": "npm ai-setup-doctor@0.1.0-beta.0 (latest+beta); hygiene suite #2 (diagnose)",
        "market": "Same AI-dev tool wave; diagnostic CLIs (doctor pattern) have high shareability.",
        "competitors": "Ad-hoc scripts, security scanners (generic), MCP inspector tools",
        "why_now": "MCP+multi-agent setup pain is acute; doctor UX is viral (npx check).",
        "risks": "Overlap with secret-guard / instruct-sync; needs clear product boundary",
        "evidence": "Fixtures for bad MCP, secrets, drift, bloat already in repo",
        "next_research": "Which checks users value most; CI integration demand",
        "mvp": "check (default), --json, --strict, --cwd; core agent/MCP/secret checks",
        "stack": "TypeScript, Node, CLI",
        "effort_weeks": 1,
        "delivery_notes": "Publish to npm; couple with instruct-sync funnel",
        "icp": "AI coding developers setting up new machines/repos",
        "pricing": "OSS free; CI cloud later",
        "gtm": "Bundle mention in instruct-sync README; Twitter demos of red/green checks",
        "revenue_model": "OSS + optional hosted CI",
        "arpu_hint": "$0–$20/mo CI tier",
    },
    {
        "id": "P03",
        "name": "mcp-sync",
        "category": "AI Tooling",
        "one_liner": "Canonical MCP server config sync across Cursor, Claude, VS Code, Windsurf clients",
        "problem": "MCP server JSON configs diverge per client; env vars and secrets get copy-pasted unsafely.",
        "burn": 4,
        "soft_native": 5,
        "ship": 5,
        "dist": 4,
        "moat": 3,
        "status": "Published",
        "stage": "Shipped / iterate",
        "platforms": "CLI,API",
        "repo": "mcp-sync",
        "tags": "AI,MCP,CLI,config,suite",
        "notes": "npm mcp-config-sync@0.2.0 + @bhaskarauthor/mcp-sync; hygiene suite #3; co-located at development/mcp-sync (own git remote)",
        "market": "MCP becoming standard transport for agent tools; multi-client users growing fast.",
        "competitors": "Manual JSON, vendor UI only, early MCP managers",
        "why_now": "MCP ecosystem still immature; window for open standard tooling.",
        "risks": "Spec churn; official multi-client managers from Anthropic/OpenAI/MS",
        "evidence": "Same distribution motion as instruct-sync; shared adapter patterns",
        "next_research": "Map all client MCP file locations & schema diffs",
        "mvp": "status/diff/sync for 4 clients; secret redaction; dry-run",
        "stack": "TypeScript CLI; shared libs with instruct-sync",
        "effort_weeks": 3,
        "delivery_notes": "Shipped; grow adoption alongside instruct-sync funnel",
        "icp": "Power users of MCP across multiple IDEs",
        "pricing": "OSS; Pro for org-wide MCP catalogs",
        "gtm": "MCP community, Discord, HN, couple with instruct-sync",
        "revenue_model": "Open core",
        "arpu_hint": "$12/seat org catalog",
    },
    {
        "id": "P04",
        "name": "secret-guard",
        "category": "AI Tooling / Security",
        "one_liner": "Pre-commit and CI guard against secrets leaking into AI agent rules, MCP env, and prompts",
        "problem": "API keys land in .cursor, CLAUDE.md, MCP env blocks, and chat exports.",
        "burn": 5,
        "soft_native": 5,
        "ship": 4,
        "dist": 4,
        "moat": 3,
        "status": "Building",
        "stage": "MVP CLI",
        "platforms": "CLI,Extension,API",
        "repo": "secret-guard",
        "tags": "security,secrets,AI,CI",
        "notes": "Local package ai-secret-guard@0.1.0-beta.0; scan AI paths + MCP env + SARIF; co-located at development/secret-guard",
        "market": "Secrets sprawl + AI configs = new leak surface; security budgets growing.",
        "competitors": "gitleaks, trufflehog, GitHub secret scanning (not AI-path aware)",
        "why_now": "AI config files are new secret vectors poorly covered by classic scanners.",
        "risks": "Crowded secrets market; need AI-path differentiation",
        "evidence": "setup-doctor leaky-secrets fixtures; enterprise pain stories",
        "next_research": "False positive rates; VS Code extension demand",
        "mvp": "Scan agent dirs + MCP JSON; pre-commit hook; SARIF output",
        "stack": "Go or Rust scanner core + TS wrapper; or pure TS",
        "effort_weeks": 4,
        "delivery_notes": "Start as setup-doctor module; split when traction",
        "icp": "Security-conscious teams using AI coding tools",
        "pricing": "OSS CLI; team dashboard $15–40/dev/mo",
        "gtm": "Security Twitter, DevSecOps blogs, GitHub Action marketplace",
        "revenue_model": "Open core SaaS",
        "arpu_hint": "$25/seat team",
    },
    {
        "id": "P05",
        "name": "llm-spend",
        "category": "AI Tooling / FinOps",
        "one_liner": "Unified LLM cost, token, and budget observability across OpenAI, Anthropic, Google, Azure, open models",
        "problem": "Teams burn LLM budget with zero allocation visibility by project/user/feature.",
        "burn": 5,
        "soft_native": 5,
        "ship": 3,
        "dist": 3,
        "moat": 3,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,API,CLI",
        "repo": "llm-spend",
        "tags": "AI,FinOps,observability,cost",
        "notes": "Superseded for NOW by P30 agent-spend-guard (guardrails wedge); keep as broad-platform LATER path",
        "market": "LLM spend becoming material OpEx; CFOs ask for unit economics of AI features.",
        "competitors": "Helicone, Langfuse, Portkey, vendor dashboards, CloudZero AI modules",
        "why_now": "Multi-provider reality; startups need budget caps before Series A burns.",
        "risks": "Crowded observability; need SMB wedge or India/SME pricing",
        "evidence": "Surveys of runaway GPT bills; proxy open-source traction",
        "next_research": "Willingness to pay SMB vs enterprise; proxy vs pull-billing UX",
        "mvp": "Proxy + per-key budgets + simple dashboard + alerts",
        "stack": "TS (Hono/Fastify), Postgres, React, provider SDKs",
        "effort_weeks": 8,
        "delivery_notes": "Larger surface; not NOW unless paid design partner",
        "icp": "Startups with multi-LLM apps; AI platform teams",
        "pricing": "Free tier; $49–299/mo by volume",
        "gtm": "Product Hunt, AI eng communities, content on unit cost",
        "revenue_model": "SaaS usage tiers",
        "arpu_hint": "$99/mo mid",
    },
    {
        "id": "P06",
        "name": "skill-sync",
        "category": "AI Tooling",
        "one_liner": "Package, version, and sync agent skills/tools/prompts like packages across teams",
        "problem": "Custom agent skills and prompt packs are copy-pasted Slackware with no versions.",
        "burn": 3,
        "soft_native": 5,
        "ship": 3,
        "dist": 3,
        "moat": 4,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "CLI,API,Web",
        "repo": "skill-sync",
        "tags": "AI,skills,registry,devtools",
        "notes": "Depends on skills format maturity (Claude skills, custom tools)",
        "market": "Agent skills emerging as unit of reuse; registries will form.",
        "competitors": "Internal wikis, prompt marketplaces, LangChain hubs",
        "why_now": "Early—format still fluid; park until standards stabilize slightly.",
        "risks": "Premature; vendors lock skills into their ecosystems",
        "evidence": "Analogous to npm for agent capabilities narrative",
        "next_research": "Track Claude/Cursor skill formats quarterly",
        "mvp": "Skill package schema + install/publish CLI + private registry",
        "stack": "TS CLI + simple registry (S3 + metadata DB)",
        "effort_weeks": 10,
        "delivery_notes": "LATER after instruct/mcp patterns proven",
        "icp": "Platform teams standardizing agent capabilities",
        "pricing": "Free public; private registry $99+/mo",
        "gtm": "DevRel, skill examples, partner agents",
        "revenue_model": "Registry SaaS",
        "arpu_hint": "$150/mo org",
    },
    {
        "id": "P07",
        "name": "shadow-AI",
        "category": "AI Tooling / Security / GRC",
        "one_liner": "Discover unauthorized AI tools, browser extensions, and SaaS usage across the enterprise",
        "problem": "Employees paste company data into consumer ChatGPT/Claude without IT visibility.",
        "burn": 5,
        "soft_native": 4,
        "ship": 2,
        "dist": 2,
        "moat": 3,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,Extension,API",
        "repo": "shadow-ai",
        "tags": "security,GRC,CASB,AI",
        "notes": "Enterprise sales motion; not indie-friendly",
        "market": "Shadow IT + GenAI = board-level risk; CASB/SSPM vendors expanding.",
        "competitors": "Netskope, Zscaler, Microsoft Purview, specialized AI governance startups",
        "why_now": "Regulators and CISOs waking up; budgets exist but sales cycles long.",
        "risks": "Enterprise sales, privacy, needs endpoint/network hooks",
        "evidence": "Gartner-style AI governance category forming",
        "next_research": "SMB browser-extension wedge vs full CASB",
        "mvp": "Browser extension + domain allowlist + admin dashboard (pilot)",
        "stack": "Chrome ext, Node API, React admin",
        "effort_weeks": 12,
        "delivery_notes": "PARK for solo; partner or later",
        "icp": "CISOs at mid-market / regulated industries",
        "pricing": "Per seat $5–15/mo enterprise",
        "gtm": "Security conferences, MSSP partners, content on data leakage",
        "revenue_model": "Enterprise SaaS",
        "arpu_hint": "$8/seat",
    },
    {
        "id": "P08",
        "name": "eval-harness",
        "category": "AI Tooling",
        "one_liner": "Lightweight eval harness for prompt/agent regression tests in CI for product teams",
        "problem": "Prompt changes break behavior silently; full ML platforms are overkill for app teams.",
        "burn": 4,
        "soft_native": 5,
        "ship": 3,
        "dist": 3,
        "moat": 3,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "CLI,API,Web",
        "repo": "eval-harness",
        "tags": "AI,eval,testing,CI",
        "notes": "Compete on simplicity vs LangSmith/Promptfoo",
        "market": "Eval tooling category heating; every serious AI product needs regression.",
        "competitors": "Promptfoo, LangSmith, Braintrust, DeepEval, custom pytest",
        "why_now": "Agent apps in production need CI; many teams still ad-hoc.",
        "risks": "Crowded; hard to differentiate without unique datasets or UX",
        "evidence": "Promptfoo OSS traction proves demand",
        "next_research": "Niche: India cost-sensitive or agent-trace specific",
        "mvp": "YAML cases, multi-provider run, score thresholds, GitHub Action",
        "stack": "Python or TS CLI + optional dashboard",
        "effort_weeks": 6,
        "delivery_notes": "NEXT if design partner; else watch",
        "icp": "Product engineers shipping LLM features",
        "pricing": "OSS; cloud runs $29–199/mo",
        "gtm": "GitHub Action + docs SEO + AI eng Discord",
        "revenue_model": "Open core SaaS",
        "arpu_hint": "$79/mo",
    },
    {
        "id": "P09",
        "name": "auth-anomaly-radar",
        "category": "Security / Identity / Fraud",
        "one_liner": "Lightweight auth anomaly detection for SaaS (impossible travel, credential stuffing signals)",
        "problem": "SMBs enable Auth0/Cognito but lack fraud signals without enterprise SIEM.",
        "burn": 4,
        "soft_native": 5,
        "ship": 3,
        "dist": 2,
        "moat": 3,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "API,Web",
        "repo": "auth-anomaly-radar",
        "tags": "security,identity,fraud,SaaS",
        "notes": "Integrate as webhook consumer for auth providers",
        "market": "Identity attacks rising; mid-market cannot afford full fraud stacks.",
        "competitors": "Arkose, Sift, Fingerprint, Auth0 Attack Protection, Cloudflare",
        "why_now": "Credential stuffing commoditized; need affordable SMB layer.",
        "risks": "Incumbents bundle features; false positives hurt UX",
        "evidence": "Auth provider event APIs enable third-party detection",
        "next_research": "WTP for $99–499/mo vs free Auth0 features",
        "mvp": "Ingest login events → risk score → Slack/webhook",
        "stack": "Python/Go service, Postgres, React",
        "effort_weeks": 10,
        "delivery_notes": "LATER; needs trust + data",
        "icp": "SaaS companies 10–200 eng, India SaaS exporters",
        "pricing": "MAU-based $99–999/mo",
        "gtm": "Auth0/Clerk marketplaces, security newsletters",
        "revenue_model": "SaaS",
        "arpu_hint": "$250/mo",
    },
    {
        "id": "P10",
        "name": "fraud-signal-kit",
        "category": "Security / Identity / Fraud",
        "one_liner": "Composable fraud signal APIs (device, velocity, email risk) for Indian fintech and marketplaces",
        "problem": "India digital payments fraud needs local signal stacks cheaper than global suites.",
        "burn": 4,
        "soft_native": 4,
        "ship": 2,
        "dist": 2,
        "moat": 3,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "API",
        "repo": "fraud-signal-kit",
        "tags": "fraud,fintech,India,API",
        "notes": "Regulatory/compliance heavy; partner path",
        "market": "India UPI/e-com fraud scale; local players + banks buying signals.",
        "competitors": "Signzy, Bureau, TransUnion, global Device Intelligence",
        "why_now": "Digital public infrastructure + fintech boom, but hard regulated moat.",
        "risks": "Data partnerships, capital, regulation",
        "evidence": "India fintech security spend growing with volume",
        "next_research": "Partnership with KYC vendors; niche vertical (BNPL, gaming)",
        "mvp": "3 signals (email, velocity, device fingerprint lite) + sandbox",
        "stack": "Go API, Redis, Postgres",
        "effort_weeks": 16,
        "delivery_notes": "PARK solo without partners",
        "icp": "India fintech / marketplace risk teams",
        "pricing": "Per API call + MRR",
        "gtm": "Fintech meetups, RBI ecosystem events, founders network",
        "revenue_model": "Usage API",
        "arpu_hint": "Usage-variable",
    },
    {
        "id": "P11",
        "name": "cloud-waste-radar",
        "category": "Cloud FinOps",
        "one_liner": "Find idle cloud resources and rightsizing opportunities for AWS/GCP/Azure startups",
        "problem": "Surveys classically cite ~30% cloud waste; startups overprovision without FinOps staff.",
        "burn": 4,
        "soft_native": 5,
        "ship": 3,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,CLI,API",
        "repo": "cloud-waste-radar",
        "tags": "FinOps,cloud,AWS,cost",
        "notes": "Crowded but SMB pricing gap exists",
        "market": "Cloud FinOps large; waste ~30% class from industry surveys (Flexera et al.).",
        "competitors": "Vantage, CloudHealth, Kubecost, native AWS Cost Explorer, nOps",
        "why_now": "Interest rates / efficiency culture; AI infra spend spike.",
        "risks": "Commoditization; AWS native improvements",
        "evidence": "Always-on demand for idle EBS/EIP/RDS detection",
        "next_research": "India price-sensitive wedge; agentic auto-remediation angle",
        "mvp": "AWS read-only connector + idle resource report + Slack digest",
        "stack": "Python, AWS SDK, Next.js",
        "effort_weeks": 8,
        "delivery_notes": "LATER unless clear niche",
        "icp": "Startups $5k–50k/mo cloud bill without FinOps hire",
        "pricing": "% of savings or $99–499/mo",
        "gtm": "Content on waste, Product Hunt, AWS Marketplace",
        "revenue_model": "SaaS + savings share",
        "arpu_hint": "$199/mo",
    },
    {
        "id": "P12",
        "name": "dev-onboard-os",
        "category": "Dev Productivity",
        "one_liner": "Opinionated new-hire engineering onboarding OS: env, access, docs, first PR in days not weeks",
        "problem": "Onboarding still tribal; new engineers lose 2–4 weeks to setup and tribal knowledge.",
        "burn": 3,
        "soft_native": 4,
        "ship": 3,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,CLI",
        "repo": "dev-onboard-os",
        "tags": "onboarding,devtools,productivity",
        "notes": "Overlaps DX platforms; hard differentiation",
        "market": "DevEx budgets rising; remote/hybrid makes onboarding harder.",
        "competitors": "Compass, Internal.io, Notion templates, custom scripts",
        "why_now": "AI can personalize onboarding paths; still crowded.",
        "risks": "Low willingness to pay; internal tools suffice",
        "evidence": "DX surveys show onboarding as top pain",
        "next_research": "Vertical: AI-era onboarding (agents + secrets + MCP day-1)",
        "mvp": "Checklist product + golden path CLI + progress for managers",
        "stack": "Next.js, Postgres, integrations (GitHub, Slack)",
        "effort_weeks": 10,
        "delivery_notes": "Tie to setup-doctor for AI-era wedge later",
        "icp": "Eng managers at 20–200 person product orgs",
        "pricing": "$8–15/seat/mo",
        "gtm": "Eng manager content, GitHub App",
        "revenue_model": "SaaS seats",
        "arpu_hint": "$12/seat",
    },
    {
        "id": "P13",
        "name": "sbom-lite",
        "category": "Security / Supply Chain",
        "one_liner": "Zero-friction SBOM generate, diff, and policy gate for npm/PyPI/Go in CI",
        "problem": "Executive orders and enterprise buyers demand SBOMs; SMBs find tools heavy.",
        "burn": 3,
        "soft_native": 5,
        "ship": 4,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "CLI,API",
        "repo": "sbom-lite",
        "tags": "SBOM,security,supply-chain,CI",
        "notes": "Syft/Grype ecosystem strong; need UX wedge",
        "market": "Software supply chain security mandatory for many RFPs.",
        "competitors": "Syft, Anchore, Snyk, FOSSA, Dependency-Track",
        "why_now": "Compliance pull; open standards (SPDX/CycloneDX).",
        "risks": "Incumbents free/open; hard monetization",
        "evidence": "GitHub dependency graph + customer questionnaires",
        "next_research": "Policy-as-code templates for India exporters / ISO",
        "mvp": "One-command SBOM + PR comment + fail on license/CVE policy",
        "stack": "Go CLI wrapping standards; GitHub Action",
        "effort_weeks": 6,
        "delivery_notes": "LATER OSS brand play",
        "icp": "Security engineers at product companies selling B2B",
        "pricing": "OSS; SaaS policy hub $99+/mo",
        "gtm": "GitHub Action marketplace, compliance content",
        "revenue_model": "Open core",
        "arpu_hint": "$99/mo",
    },
    {
        "id": "P14",
        "name": "grc-evidence-autopilot",
        "category": "Compliance / GRC",
        "one_liner": "Auto-collect compliance evidence (access logs, policies, screenshots) for SOC2/ISO startups",
        "problem": "SOC2 prep is spreadsheet hell; evidence collection is continuous toil.",
        "burn": 4,
        "soft_native": 4,
        "ship": 2,
        "dist": 2,
        "moat": 3,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,API",
        "repo": "grc-evidence-autopilot",
        "tags": "GRC,SOC2,compliance,SaaS",
        "notes": "Crowded (Vanta, Drata); India cost wedge possible",
        "market": "GRC automation multi-billion trajectory; every B2B SaaS needs trust.",
        "competitors": "Vanta, Drata, Secureframe, Sprinto (India), Scrut",
        "why_now": "Still expensive for early startups; AI evidence drafting angle.",
        "risks": "Well-funded incumbents; trust & auditor relationships",
        "evidence": "Sprinto shows India GTM works for GRC",
        "next_research": "Niche frameworks (DPDP India) vs SOC2 clone",
        "mvp": "Integrations (AWS, GitHub, Google) + control map + auditor share link",
        "stack": "Next.js, workers, encrypted store",
        "effort_weeks": 16,
        "delivery_notes": "PARK unless India DPDP-specific wedge",
        "icp": "Series A SaaS needing SOC2 fast; India SaaS exporters",
        "pricing": "$200–1000/mo",
        "gtm": "Auditor partnerships, YC-style communities",
        "revenue_model": "SaaS",
        "arpu_hint": "$400/mo",
    },
    {
        "id": "P15",
        "name": "wa-ops-desk",
        "category": "India SMB / WhatsApp Ops",
        "one_liner": "WhatsApp-first ops desk for Indian SMBs: orders, FAQs, staff routing, CRM lite",
        "problem": "Indian SMBs live on WhatsApp; Excel + personal phones do not scale multi-staff.",
        "burn": 5,
        "soft_native": 4,
        "ship": 3,
        "dist": 4,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,Android,API",
        "repo": "wa-ops-desk",
        "tags": "India,SMB,WhatsApp,CRM",
        "notes": "Meta BSP dependency; regulatory/pricing constraints",
        "market": "India DX ~$124–144B class 2025–26; SME digitization high CAGR; WhatsApp ubiquitous.",
        "competitors": "Interakt, Wati, AiSensy, Gallabox, custom BSP builds",
        "why_now": "Cloud API mature; SMBs will pay for shared inbox + catalog.",
        "risks": "BSP competition, Meta policy, support burden",
        "evidence": "Many funded WA SaaS; still room in vertical niches",
        "next_research": "Pick vertical (clinics, tutors, local retail) for wedge",
        "mvp": "Shared inbox + labels + quick replies + 1 catalog flow",
        "stack": "Next.js, WhatsApp Cloud API, Postgres, Razorpay",
        "effort_weeks": 10,
        "delivery_notes": "NEXT if India GTM bandwidth",
        "icp": "India SMB 2–20 staff using WhatsApp Business",
        "pricing": "₹999–4999/mo tiers",
        "gtm": "WhatsApp communities, CA/partners, local SEO, YouTube Hindi",
        "revenue_model": "SaaS INR",
        "arpu_hint": "₹2,000/mo",
    },
    {
        "id": "P16",
        "name": "gst-ops-copilot",
        "category": "India SMB / Fintech",
        "one_liner": "GST filing prep copilot: invoice hygiene, GSTR mismatch alerts, CA handoff packs",
        "problem": "SMEs struggle with invoice data quality and GSTR mismatches before CA filing.",
        "burn": 4,
        "soft_native": 4,
        "ship": 2,
        "dist": 3,
        "moat": 3,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,API",
        "repo": "gst-ops-copilot",
        "tags": "India,GST,fintech,SMB,compliance",
        "notes": "Heavy domain + API/partner dependencies",
        "market": "India GST ecosystem huge; Clear, Zoho, Tally dominate but AI copilots emerging.",
        "competitors": "ClearTax, Zoho Books, Tally, Khatabook-adjacent, CA tools",
        "why_now": "AI for document extraction; still trust/regulatory friction.",
        "risks": "Incumbents, GSTN API access, liability",
        "evidence": "Always-green tax software category in India",
        "next_research": "Partner with CAs rather than replace; API access path",
        "mvp": "Invoice OCR + mismatch checklist + export for CA",
        "stack": "Python OCR/LLM, Next.js, secure storage",
        "effort_weeks": 14,
        "delivery_notes": "LATER/domain partner",
        "icp": "India SMEs + small CA firms",
        "pricing": "₹499–2999/mo",
        "gtm": "CA channel, YouTube GST explainers",
        "revenue_model": "SaaS + CA white-label",
        "arpu_hint": "₹1,000/mo",
    },
    {
        "id": "P17",
        "name": "appt-book-india",
        "category": "India SMB / Appointments",
        "one_liner": "Lightweight appointment booking with WhatsApp reminders for salons, tutors, clinics lite",
        "problem": "No-shows and phone tag; Calendly-class tools not localized/priced for Bharat SMB.",
        "burn": 4,
        "soft_native": 4,
        "ship": 4,
        "dist": 4,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,Android,API",
        "repo": "appt-book-india",
        "tags": "India,SMB,booking,WhatsApp",
        "notes": "Can share stack with clinic/WA products",
        "market": "Local services digitizing; WhatsApp reminders reduce no-shows dramatically.",
        "competitors": "Practo (health), Fresha, Setmore, Google Booking, local apps",
        "why_now": "UPI + WA + cheap smartphones = booking UX ready.",
        "risks": "Low ARPU, support in vernacular, churn",
        "evidence": "High no-show rates; WA reminder efficacy known",
        "next_research": "Salon vs tutor vertical first; offline-first Android",
        "mvp": "Public booking page + staff calendar + WA reminder via Cloud API",
        "stack": "Next.js, Postgres, WA API, Razorpay deposits optional",
        "effort_weeks": 6,
        "delivery_notes": "NEXT India track candidate",
        "icp": "Salons, tutors, small clinics India tier-1/2",
        "pricing": "₹499–1499/mo",
        "gtm": "Local Facebook groups, partners, Instagram ads",
        "revenue_model": "SaaS INR",
        "arpu_hint": "₹799/mo",
    },
    {
        "id": "P18",
        "name": "clinic-admin-lite",
        "category": "India SMB / Health Ops",
        "one_liner": "Clinic admin lite: appointments, patient ledger, Rx print, WhatsApp follow-ups for small practices",
        "problem": "Small Indian clinics still paper/Excel; full HMS is overkill and expensive.",
        "burn": 4,
        "soft_native": 4,
        "ship": 3,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,Android",
        "repo": "clinic-admin-lite",
        "tags": "health,India,SMB,clinic",
        "notes": "Health data sensitivity; start non-diagnostic admin only",
        "market": "Fragmented outpatient market; digitization slow but inevitable.",
        "competitors": "Practo Ray, eHospital, local HMS, Excel",
        "why_now": "Post-COVID acceptance of digital; WA patient comms norm.",
        "risks": "Trust, data residency, long sales, support",
        "evidence": "Many micro-clinics underserved by enterprise HMS",
        "next_research": "Single-doctor clinic MVP interviews in 1 city",
        "mvp": "Patients + appts + basic billing + WA reminder",
        "stack": "Next.js/Flutter, Postgres, encryption at rest",
        "effort_weeks": 12,
        "delivery_notes": "LATER after appt-book learns domain",
        "icp": "1–5 doctor clinics India",
        "pricing": "₹999–3999/mo",
        "gtm": "Medical associations, CA/IT partners, referrals",
        "revenue_model": "SaaS",
        "arpu_hint": "₹1,500/mo",
    },
    {
        "id": "P19",
        "name": "learn-loop",
        "category": "Education",
        "one_liner": "Spaced-repetition + cohort homework OS for coaching institutes and online tutors",
        "problem": "Tutors lack simple tool for assignments, revision loops, and parent updates.",
        "burn": 3,
        "soft_native": 4,
        "ship": 3,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,Android,iOS",
        "repo": "learn-loop",
        "tags": "edtech,tutors,India,SRS",
        "notes": "Edtech crowded; niche tutor OS possible",
        "market": "India edtech reset favors tools over content marketplaces.",
        "competitors": "Google Classroom, Teachmint, Classplus, Anki (SRS only)",
        "why_now": "Tutors want ops not VC-scale marketplaces.",
        "risks": "Low willingness to pay; free Google tools",
        "evidence": "Teachmint-style SMB tutor tools still active",
        "next_research": "Paid tutor interviews; WA parent updates wedge",
        "mvp": "Classes + assignments + SRS deck + parent WA digest",
        "stack": "Flutter + Firebase or Next.js",
        "effort_weeks": 10,
        "delivery_notes": "PARK unless passion domain",
        "icp": "Independent tutors & small coaching centers",
        "pricing": "₹299–999/mo",
        "gtm": "Tutor communities, YouTube",
        "revenue_model": "SaaS",
        "arpu_hint": "₹499/mo",
    },
    {
        "id": "P20",
        "name": "cyber-smb-shield",
        "category": "Cyber / SMB",
        "one_liner": "Managed cybersecurity starter pack for SMBs: DNS filter, phishing drills, backup check, score",
        "problem": "SMBs are ransomware targets but cannot staff a SOC.",
        "burn": 4,
        "soft_native": 3,
        "ship": 2,
        "dist": 2,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,API",
        "repo": "cyber-smb-shield",
        "tags": "cyber,SMB,MSSP,security",
        "notes": "Services-heavy; partner with MSSPs",
        "market": "AI cyber / managed services large and growing; SMB underserved.",
        "competitors": "Huntress, CrowdStrike Falcon Go, local MSSPs, Microsoft 365 Defender",
        "why_now": "Ransomware economics + insurance questionnaires.",
        "risks": "Liability, 24/7 expectations, capital",
        "evidence": "MSSP attach rates rising globally",
        "next_research": "Software-only score vs full MDR; India pricing",
        "mvp": "Security score questionnaire + phishing sim + DNS blocklist SaaS",
        "stack": "Next.js + third-party DNS/email APIs",
        "effort_weeks": 12,
        "delivery_notes": "PARK without security partner",
        "icp": "SMB 10–100 employees, India & global",
        "pricing": "$5–15/seat/mo",
        "gtm": "MSP channel, insurance brokers",
        "revenue_model": "SaaS + channel",
        "arpu_hint": "$8/seat",
    },
    {
        "id": "P21",
        "name": "data-quality-guard",
        "category": "Data Quality",
        "one_liner": "Pipeline data quality checks as code with anomaly alerts for analytics teams",
        "problem": "Silent data breakage destroys dashboards; Great Expectations-class tools feel heavy.",
        "burn": 3,
        "soft_native": 5,
        "ship": 3,
        "dist": 2,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "CLI,API,Web",
        "repo": "data-quality-guard",
        "tags": "data,quality,analytics,observability",
        "notes": "Crowded (GE, Monte Carlo, Soda)",
        "market": "Data observability category established; mid-market gap on price.",
        "competitors": "Monte Carlo, Soda, Great Expectations, Elementary, Anomalo",
        "why_now": "AI features amplify bad data cost; still crowded.",
        "risks": "Hard to unseat; long enterprise sales",
        "evidence": "Data downtime costly anecdotes everywhere",
        "next_research": "Warehouse-native cheap tier for startups",
        "mvp": "SQL expectation YAML + dbt/Airflow hooks + Slack",
        "stack": "Python, SQL engines, Next.js light UI",
        "effort_weeks": 10,
        "delivery_notes": "PARK unless unique AI-data angle",
        "icp": "Analytics eng at Series A–C",
        "pricing": "$99–499/mo",
        "gtm": "dbt community, data Twitter",
        "revenue_model": "SaaS",
        "arpu_hint": "$199/mo",
    },
    {
        "id": "P22",
        "name": "care-companion",
        "category": "Consumer Care",
        "one_liner": "Family care coordination app: meds, appointments, shared notes for elders",
        "problem": "Adult children coordinate elder care over chaotic WhatsApp groups.",
        "burn": 3,
        "soft_native": 3,
        "ship": 3,
        "dist": 2,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "iOS,Android,Web",
        "repo": "care-companion",
        "tags": "consumer,health,family,care",
        "notes": "Emotional domain; trust & retention hard",
        "market": "Aging populations; family caregiving apps emerging globally.",
        "competitors": "Lotsa, CareZone (hist), Google Keep/WhatsApp defaults",
        "why_now": "Mobile-first families; still behavior change challenge.",
        "risks": "Consumer CAC, medical liability optics",
        "evidence": "Caregiver stress widely documented",
        "next_research": "India joint-family WA-first design interviews",
        "mvp": "Shared list + med reminders + appt calendar for 1 family",
        "stack": "Flutter, Firebase",
        "effort_weeks": 10,
        "delivery_notes": "PARK consumer",
        "icp": "Multi-city families caring for parents",
        "pricing": "Freemium; ₹99–299/mo family",
        "gtm": "Content, hospitals partners, ASO",
        "revenue_model": "Consumer sub",
        "arpu_hint": "₹149/mo",
    },
    {
        "id": "P23",
        "name": "personal-crm",
        "category": "Personal CRM",
        "one_liner": "Privacy-first personal CRM for founders: follow-ups, context notes, birthday nudges",
        "problem": "LinkedIn is not a relationship system; people forget follow-ups.",
        "burn": 3,
        "soft_native": 4,
        "ship": 4,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,iOS,Desktop",
        "repo": "personal-crm",
        "tags": "CRM,productivity,personal,privacy",
        "notes": "Many indie apps; differentiation via privacy + AI summarize",
        "market": "Indie personal CRM niche durable (Clay-like personal, Monica, etc.).",
        "competitors": "Monica, Dex, Clay, Notion templates, Apple Contacts",
        "why_now": "AI can draft follow-ups; privacy concerns rising.",
        "risks": "Habit formation; low willingness to pay",
        "evidence": "Recurring indie launches with modest paid bases",
        "next_research": "Local-first encrypted MVP interest",
        "mvp": "People + interactions + remind + import contacts",
        "stack": "Local-first (SQLite) + optional sync; Tauri/Next",
        "effort_weeks": 6,
        "delivery_notes": "Side project LATER",
        "icp": "Founders, consultants, community builders",
        "pricing": "$8–12/mo",
        "gtm": "Indie Hackers, Product Hunt, Twitter",
        "revenue_model": "Consumer SaaS",
        "arpu_hint": "$10/mo",
    },
    {
        "id": "P24",
        "name": "creator-ops",
        "category": "Creator Ops",
        "one_liner": "Ops OS for mid-tier creators: content calendar, sponsor CRM, asset vault, analytics rollup",
        "problem": "Creators juggle Notion + Sheets + Drive + native analytics without an ops spine.",
        "burn": 3,
        "soft_native": 4,
        "ship": 3,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web",
        "repo": "creator-ops",
        "tags": "creator,ops,media,SaaS",
        "notes": "Crowded creator tooling",
        "market": "Creator economy large; tools consolidating.",
        "competitors": "Later, Notion, Hoppycopy, sponsor marketplaces, native suites",
        "why_now": "Multi-platform creators need ops not just scheduling.",
        "risks": "Platform API limits; fashion-driven churn",
        "evidence": "Agencies pay for ops; solo creators price-sensitive",
        "next_research": "Agency multi-client mode first",
        "mvp": "Calendar + sponsor pipeline + asset links",
        "stack": "Next.js, Postgres",
        "effort_weeks": 8,
        "delivery_notes": "LATER",
        "icp": "Creators 50k–500k followers or small agencies",
        "pricing": "$19–79/mo",
        "gtm": "YouTube/Twitter creators, affiliate",
        "revenue_model": "SaaS",
        "arpu_hint": "$29/mo",
    },
    {
        "id": "P25",
        "name": "climate-ops-meter",
        "category": "Climate / ESG",
        "one_liner": "Lightweight Scope 1–3 carbon estimate and supplier survey tool for mid-market exporters",
        "problem": "Exporters face ESG questionnaires without enterprise sustainability teams.",
        "burn": 3,
        "soft_native": 4,
        "ship": 2,
        "dist": 2,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web",
        "repo": "climate-ops-meter",
        "tags": "climate,ESG,carbon,compliance",
        "notes": "Methodology trust critical",
        "market": "Carbon accounting software growing with regulation (EU CSRD ripple).",
        "competitors": "Watershed, Persefoni, local consultancies, Excel",
        "why_now": "Supply-chain pressure on Indian manufacturers/exporters.",
        "risks": "Methodology disputes; long sales; consulting hybrid",
        "evidence": "Buyer questionnaires proliferating",
        "next_research": "India exporter pilot with one industry body",
        "mvp": "Activity data forms + emission factors + PDF report",
        "stack": "Next.js, Postgres, factor DB",
        "effort_weeks": 12,
        "delivery_notes": "PARK without domain partner",
        "icp": "India mid-market exporters selling to EU/US",
        "pricing": "$199–999/mo",
        "gtm": "Trade associations, consultants",
        "revenue_model": "SaaS + services",
        "arpu_hint": "$400/mo",
    },
    {
        "id": "P26",
        "name": "supply-chain-visibility-lite",
        "category": "Supply Chain",
        "one_liner": "Lite multi-tier supplier status & risk notes for SMEs without SAP budgets",
        "problem": "SMEs lack visibility beyond tier-1 suppliers; enterprise SC tools are overkill.",
        "burn": 3,
        "soft_native": 3,
        "ship": 2,
        "dist": 2,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Web,API",
        "repo": "sc-visibility-lite",
        "tags": "supply-chain,SME,risk",
        "notes": "Data acquisition is the product",
        "market": "Supply chain risk software large enterprise; SME gap remains.",
        "competitors": "Resilinc, SAP Ariba, Excel + email",
        "why_now": "Geopolitical risk awareness; still hard data problem.",
        "risks": "Chicken-egg supplier data; enterprise sales DNA",
        "evidence": "Post-pandemic SC investment continued",
        "next_research": "Narrow vertical (textiles/auto components India)",
        "mvp": "Supplier directory + status check-ins + risk tags",
        "stack": "Next.js, Postgres, WhatsApp check-in bots",
        "effort_weeks": 14,
        "delivery_notes": "PARK",
        "icp": "India manufacturing SMEs in export chains",
        "pricing": "₹5,000–25,000/mo",
        "gtm": "Industry associations",
        "revenue_model": "SaaS",
        "arpu_hint": "₹10,000/mo",
    },
    {
        "id": "P27",
        "name": "focus-forge",
        "category": "Focus / Productivity Apps",
        "one_liner": "Deep-work OS: site blocker + session goals + AI weekly review for knowledge workers",
        "problem": "Attention fragmentation; existing blockers lack outcome review loops.",
        "burn": 3,
        "soft_native": 4,
        "ship": 4,
        "dist": 3,
        "moat": 1,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "Desktop,Extension,iOS",
        "repo": "focus-forge",
        "tags": "productivity,focus,consumer",
        "notes": "App store graveyard risk; lifestyle product",
        "market": "Huge consumer productivity market; retention brutal.",
        "competitors": "Freedom, Cold Turkey, Opal, Forest, Notion habits",
        "why_now": "AI review angle; still low moat.",
        "risks": "Commoditized, churn, platform policies",
        "evidence": "Always demand for blockers; hard business",
        "next_research": "Only if personal passion; else skip",
        "mvp": "Browser extension blocker + session log + weekly AI recap",
        "stack": "Chrome ext + optional Electron",
        "effort_weeks": 5,
        "delivery_notes": "PARK",
        "icp": "Knowledge workers, students",
        "pricing": "$4–8/mo",
        "gtm": "Product Hunt, TikTok productivity niche",
        "revenue_model": "Consumer sub",
        "arpu_hint": "$6/mo",
    },
    {
        "id": "P28",
        "name": "api-contract-sentinel",
        "category": "Dev Productivity / Quality",
        "one_liner": "Detect breaking API contract changes in PRs with consumer impact maps",
        "problem": "Microservices break silent consumers; OpenAPI drift is common.",
        "burn": 3,
        "soft_native": 5,
        "ship": 3,
        "dist": 3,
        "moat": 2,
        "status": "Scaffolded MVP",
        "stage": "Offline MVP + litmus",
        "platforms": "CLI,API,Extension",
        "repo": "api-contract-sentinel",
        "tags": "API,OpenAPI,CI,devtools",
        "notes": "Overlaps Optic, SwaggerHub, Spectral",
        "market": "API governance steady; platform teams buy.",
        "competitors": "Optic, Stoplight, Spectral, Redocly, Apidog",
        "why_now": "AI-generated APIs increase drift risk—possible wedge.",
        "risks": "Crowded open-source; monetization thin",
        "evidence": "PR-time contract checks valued in platform eng",
        "next_research": "AI-generated OpenAPI validation niche",
        "mvp": "OpenAPI diff in GitHub Action + break level + owners notify",
        "stack": "TypeScript/Go, GitHub App",
        "effort_weeks": 6,
        "delivery_notes": "LATER OSS brand",
        "icp": "Platform engineering teams",
        "pricing": "OSS; org $99+/mo",
        "gtm": "GitHub marketplace, platform eng blogs",
        "revenue_model": "Open core",
        "arpu_hint": "$99/mo",
    },
    {
        "id": "P29",
        "name": "agent-skill-scan",
        "category": "AI Tooling / Security",
        "one_liner": "One-command security audit of installed agent skills, MCP servers, and rules files (prompt injection, malicious payloads, secret exfil)",
        "problem": "Marketplace agent skills/MCP servers ship malicious payloads; audits found 36.8% of skills flawed and 71% of MCP servers graded F.",
        "burn": 5,
        "soft_native": 5,
        "ship": 5,
        "dist": 5,
        "moat": 3,
        "status": "Concept",
        "stage": "Design / build next",
        "platforms": "CLI,API",
        "repo": "agent-skill-scan",
        "tags": "AI,security,MCP,skills,supply-chain,suite",
        "notes": "Hygiene suite #4 (secure); rides ToxicSkills wave; reuses setup-doctor config parsers",
        "market": "Agent supply-chain security exploding: Snyk ToxicSkills found 1,467 malicious skill payloads (Feb 2026); only 29% of orgs feel prepared to secure agentic AI.",
        "competitors": "Snyk agent-scan (enterprise-flavored), mcp-scan, Cisco mcp-scanner; no indie npx-first scanner",
        "why_now": "Fresh incident data (ToxicSkills, MCP F-grades) gives a free launch narrative; fear-driven shareability; window before incumbents move downmarket.",
        "risks": "Signature arms race; Snyk/vendors ship free equivalents; false-positive trust erosion",
        "evidence": "ToxicSkills: 13.4% of 3,984 skills critical; MCP audit 71% F / 0% A; setup-doctor already parses agent configs",
        "next_research": "Injection heuristic precision/recall on real skill corpus; CI action demand",
        "mvp": "npx scan of skills/MCP/rules dirs; injection + exfil heuristics; severity report; --json/SARIF",
        "stack": "TypeScript, Node 20+, zero deps; shared parsers with setup-doctor",
        "effort_weeks": 3,
        "delivery_notes": "Build from setup-doctor secret-check module; ship while the wave is hot; absorbs P04 scope",
        "icp": "AI-assisted developers; security-conscious teams adopting agents",
        "pricing": "OSS CLI free; paid GitHub Action / org policy $15–40/dev/mo",
        "gtm": "Show HN with ToxicSkills hook, security Twitter, GitHub Action marketplace, cross-link suite READMEs",
        "revenue_model": "Open core SaaS",
        "arpu_hint": "$25/seat",
    },
    {
        "id": "P30",
        "name": "agent-spend-guard",
        "category": "AI Tooling / FinOps",
        "one_liner": "Per-dev, per-project token budgets with hard kill-switch for Claude Code, Cursor, and API agents",
        "problem": "Teams blow annual AI token budgets in months (Uber by April 2026); dashboards report after the fact, nothing enforces.",
        "burn": 5,
        "soft_native": 5,
        "ship": 4,
        "dist": 4,
        "moat": 3,
        "status": "Concept",
        "stage": "Design",
        "platforms": "CLI,API,Web",
        "repo": "agent-spend-guard",
        "tags": "AI,FinOps,budget,guardrails,tokens,suite",
        "notes": "Hygiene suite #5 (control spend); sharper guardrails wedge that supersedes P05 for NOW",
        "market": "Token bill panic is existential (FinOps Foundation: teams 3x over 2026 budget by April); 85% of GenAI deployments run without observability.",
        "competitors": "Vantage, Finout, Langfuse, LiteLLM proxy — all tracking/dashboards, none indie-sized enforcement",
        "why_now": "Q2 2026 shift from 'go fast' to 'we need guardrails, how do we control this' (FinOps Foundation, TechCrunch Jun 2026).",
        "risks": "Anthropic/Cursor ship native budget caps; proxy trust barrier; per-seat WTP unproven",
        "evidence": "TechCrunch token-bill reporting; Microsoft revoking Claude Code licenses; P05 research base",
        "next_research": "Local wrapper vs proxy UX; eng-manager WTP interviews; which agents expose usage hooks",
        "mvp": "Local usage meter + budget config + warn/stop thresholds + Slack alert; Claude Code & Cursor first",
        "stack": "TypeScript CLI + lightweight proxy; optional dashboard later",
        "effort_weeks": 4,
        "delivery_notes": "The money product of the suite; free solo meter drives paid team budgets",
        "icp": "Eng managers with runaway agent bills; indie devs on usage pricing",
        "pricing": "Free solo meter; $10–15/dev/mo team budgets + policy",
        "gtm": "Show HN, FinOps community, 'my token bill' content wave, suite cross-links",
        "revenue_model": "Open core → seat SaaS",
        "arpu_hint": "$12/seat",
    },
    {
        "id": "P31",
        "name": "indic-voice-flow",
        "category": "Consumer AI / Voice",
        "one_liner": "Wispr Flow-class voice dictation for Hinglish and Indic languages — desktop hotkey + Android keyboard",
        "problem": "English-first dictation fails code-switched Hinglish/Indic scripts; Indian professionals type far slower than they speak.",
        "burn": 4,
        "soft_native": 4,
        "ship": 2,
        "dist": 3,
        "moat": 3,
        "status": "Concept",
        "stage": "Research",
        "platforms": "Desktop,Android,iOS",
        "repo": "indic-voice-flow",
        "tags": "voice,India,consumer,dictation,ASR",
        "notes": "The Wispr-Flow-style revenue play; don't compete in English — own Hinglish. Needs cofounder or dedicated cycle.",
        "market": "Wispr Flow at ~$10M ARR / $700M valuation validates paid dictation; Indic/Hinglish segment underserved and huge.",
        "competitors": "Wispr Flow (English-first, $315M raised), Google voice typing (free), local IME apps",
        "why_now": "ASR quality for code-switched speech crossed the usable threshold; willingness to pay proven by Wispr.",
        "risks": "Google free tier; consumer CAC/churn; ASR compute cost; big departure from zero-dep CLI muscle",
        "evidence": "Wispr revenue/valuation; Whisper-class models handle Hinglish decently in 2026",
        "next_research": "On-device vs cloud ASR unit economics; 20 interviews with lawyers/doctors/CAs; ₹ WTP",
        "mvp": "Mac/Windows dictation hotkey + LLM Hinglish cleanup; Android keyboard phase 2",
        "stack": "Whisper-class ASR + LLM post-processing; Swift/Kotlin shells",
        "effort_weeks": 12,
        "delivery_notes": "NEXT, not NOW: park until suite compounds or a cofounder owns it; closed source",
        "icp": "Indian professionals (legal, medical, CA, sales) typing 2+ hrs/day",
        "pricing": "₹299–499/mo",
        "gtm": "Hindi YouTube/Instagram demos, professional associations, WhatsApp forwards",
        "revenue_model": "Consumer subscription",
        "arpu_hint": "₹399/mo",
    },
    {
        "id": "P32",
        "name": "pr-triage",
        "category": "Dev Productivity / Quality",
        "one_liner": "Risk-rank AI-generated PR diffs so humans review the 20% that actually matters",
        "problem": "Developers now spend 11.4 hrs/week reviewing AI code vs 9.8 writing; 66% cite 'almost right, but not quite' as top frustration.",
        "burn": 5,
        "soft_native": 5,
        "ship": 3,
        "dist": 3,
        "moat": 2,
        "status": "Concept",
        "stage": "Watch",
        "platforms": "CLI,API,Extension",
        "repo": "pr-triage",
        "tags": "AI,code-review,CI,devtools",
        "notes": "Big pain but CodeRabbit/Greptile/GitHub converging; watch for a hunk-level triage wedge",
        "market": "Review burden reversed the 2024 pattern; trust in AI output fell to 29% — triage is the bottleneck, not generation.",
        "competitors": "CodeRabbit, Greptile, GitHub Copilot code review, Graphite — all full reviewers, none pure triage",
        "why_now": "Review fatigue is a top-3 2026 developer pain and still underserved as a ranking (not reviewing) problem.",
        "risks": "Well-funded incumbents add triage as a feature; thin moat",
        "evidence": "2026 dev pain surveys (11.4h review, 75% more logic errors in AI PRs)",
        "next_research": "Hunk-level risk scoring accuracy; whether teams want ranking vs another reviewer",
        "mvp": "CLI + GitHub Action: rank hunks by blast radius, coverage gaps, AI-origin heuristics",
        "stack": "TypeScript + AST diff analysis; optional LLM assist",
        "effort_weeks": 6,
        "delivery_notes": "Watch; build only if P29/P30 traction frees capacity",
        "icp": "Teams merging heavy AI-generated code",
        "pricing": "OSS; org tier $99+/mo",
        "gtm": "GitHub marketplace, review-fatigue content",
        "revenue_model": "Open core",
        "arpu_hint": "$99/mo",
    },
]

assert len(OPPS) == 32, len(OPPS)

# ---------------------------------------------------------------------------
# Visibility strategy: what goes public vs stays private, and the paid layer
# OPEN-CORE    = public OSS repo (distribution/brand); money on a paid layer on top
# PRIVATE-PAID = closed source; the product itself is the paid thing
# ---------------------------------------------------------------------------
VISIBILITY = {
    "P01": ("OPEN-CORE", "Pro team sync + policy ($8–15/dev/mo)"),
    "P02": ("OPEN-CORE", "Hosted CI checks tier"),
    "P03": ("OPEN-CORE", "Org-wide MCP catalog ($12/seat)"),
    "P04": ("OPEN-CORE", "Team dashboard + SARIF CI ($15–40/dev/mo)"),
    "P05": ("PRIVATE-PAID", "SaaS usage tiers ($49–299/mo)"),
    "P06": ("OPEN-CORE", "Private registry ($99+/mo)"),
    "P07": ("PRIVATE-PAID", "Enterprise per-seat SaaS"),
    "P08": ("OPEN-CORE", "Cloud runs ($29–199/mo)"),
    "P09": ("PRIVATE-PAID", "MAU-based SaaS"),
    "P10": ("PRIVATE-PAID", "Usage-priced API"),
    "P11": ("PRIVATE-PAID", "SaaS + savings share"),
    "P12": ("PRIVATE-PAID", "Seat SaaS"),
    "P13": ("OPEN-CORE", "Policy hub SaaS ($99+/mo)"),
    "P14": ("PRIVATE-PAID", "SaaS"),
    "P15": ("PRIVATE-PAID", "INR SaaS tiers"),
    "P16": ("PRIVATE-PAID", "SaaS + CA white-label"),
    "P17": ("PRIVATE-PAID", "INR SaaS"),
    "P18": ("PRIVATE-PAID", "SaaS"),
    "P19": ("PRIVATE-PAID", "SaaS"),
    "P20": ("PRIVATE-PAID", "SaaS + MSSP channel"),
    "P21": ("PRIVATE-PAID", "SaaS"),
    "P22": ("PRIVATE-PAID", "Consumer subscription"),
    "P23": ("PRIVATE-PAID", "Consumer SaaS"),
    "P24": ("PRIVATE-PAID", "SaaS"),
    "P25": ("PRIVATE-PAID", "SaaS + services"),
    "P26": ("PRIVATE-PAID", "SaaS"),
    "P27": ("PRIVATE-PAID", "Consumer subscription"),
    "P28": ("OPEN-CORE", "Org tier ($99+/mo)"),
    "P29": ("OPEN-CORE", "Paid GitHub Action + org policy ($15–40/dev/mo)"),
    "P30": ("OPEN-CORE", "Team budgets + policy ($10–15/dev/mo)"),
    "P31": ("PRIVATE-PAID", "Consumer subscription (₹299–499/mo)"),
    "P32": ("OPEN-CORE", "Org tier ($99+/mo)"),
}
assert set(VISIBILITY) == {o["id"] for o in OPPS}


def visibility_of(o: dict) -> tuple[str, str]:
    return VISIBILITY[o["id"]]


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------

def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("00_README", 0)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 3, "B": 28, "C": 90, "D": 24})

    ws.merge_cells("B2:D2")
    ws["B2"] = "Software Opportunity Research Portfolio"
    ws["B2"].font = TITLE_FONT

    ws.merge_cells("B3:D3")
    ws["B3"] = f"Generated {TODAY} · {len(OPPS)} opportunities (P01–P{len(OPPS)}) · Scores editable · Totals & Priority are Excel formulas"
    ws["B3"].font = Font(name=FONT_NAME, size=10, italic=True, color="666666")

    sections = [
        ("How to use", [
            "1. Edit only yellow cells with blue font (score inputs 1–5) on 01_Master_Portfolio.",
            "2. TOTAL (column K) and Priority (column L) recalculate automatically—do not hardcode them.",
            "3. Priority bands: NOW ≥ 20 · NEXT 16–19 · LATER 12–15 · PARK < 12 (sum of five scores).",
            "4. Review 02–04 for research, build, and monetization detail per opportunity.",
            "5. Use 05_Scoring_Matrix for visual ranking; 06_Pipeline for phase planning.",
            "6. 07_Market_Context holds macro notes (directional figures with sources).",
            "7. Log new research in 08_Research_Log. Keep CSV export in sync when scores change (re-run generator or copy Master).",
            "8. Conventions: blue font + yellow fill = inputs; black = labels/formulas (financial modeling style).",
            "9. Visibility marker (Master cols M–N): OPEN-CORE = public OSS repo with a paid layer on top; PRIVATE-PAID = closed source, the product itself is paid.",
        ]),
        ("Sheets", [
            "00_README — this guide",
            "01_Master_Portfolio — all opportunities, scores, TOTAL formula, Priority formula",
            "02_Deep_Research — market, competitors, risks, evidence, next research",
            "03_Build_Delivery — platforms, MVP, stack, repo, effort",
            "04_Monetization — ICP, pricing, GTM, revenue model",
            "05_Scoring_Matrix — linked scores + color scale + bar chart",
            "06_Pipeline — NOW / NEXT / LATER / PARK phases",
            "07_Market_Context — macro market notes & sources",
            "08_Research_Log — dated research log entries",
        ]),
        ("Scoring dimensions (1–5)", [
            "Burn — how painful/urgent is the problem for buyers?",
            "Software-native — can this be pure software with leverage (not services-heavy)?",
            "Ship speed — how fast can a small team ship a credible MVP?",
            "Distribution — can you reach users cheaply (npx, communities, PLG)?",
            "Moat — defensibility over 2–3 years (data, network, workflow lock-in, brand).",
        ]),
        ("Files", [
            f"Excel: {XLSX_PATH.name}",
            f"CSV mirror of Master: {CSV_PATH.name}",
            f"Index: {INDEX_PATH.name}",
            "Regenerate: python3 scripts/generate_software_portfolio.py",
        ]),
        ("Working context", [
            "Shipped on npm: ai-instruct-sync@0.2.0-beta.0 (P01), ai-setup-doctor@0.1.0-beta.0 (P02), @bhaskarauthor/mcp-sync@0.2.0 + mcp-config-sync (P03).",
            "Strategic line — AI dev hygiene suite: sync (P01/P03) → diagnose (P02) → secure (P29) → control spend (P30); same audience, same npx flywheel, paid layers on top.",
            "Default bias: ship AI CLI tooling NOW; India SMB and heavier SaaS NEXT/LATER; P31 indic-voice-flow is the standalone consumer revenue bet.",
            "Figures in Market Context are directional mid-decade estimates—validate before fundraising decks.",
        ]),
    ]

    r = 5
    for title, lines in sections:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cell = ws.cell(row=r, column=2, value=title)
        cell.font = SECTION_FONT
        cell.fill = FILL_SECTION
        ws.cell(row=r, column=3).fill = FILL_SECTION
        r += 1
        for line in lines:
            ws.cell(row=r, column=2, value="•").font = BLACK
            c = ws.cell(row=r, column=3, value=line)
            c.font = BLACK
            c.alignment = WRAP
            r += 1
        r += 1

    ws.row_dimensions[2].height = 24
    for i in range(5, r):
        ws.row_dimensions[i].height = 18


def build_master(wb: Workbook) -> None:
    ws = wb.create_sheet("01_Master_Portfolio", 1)

    headers = [
        "ID", "Name", "Category", "One-liner", "Problem",
        "Burn", "Software-native", "Ship speed", "Distribution", "Moat",
        "TOTAL", "Priority", "Visibility", "Paid layer", "Status", "Stage", "Platforms", "Repo", "Tags", "Notes",
        "Last reviewed",
    ]
    # A..U

    ws["A1"] = "Software Opportunity Master Portfolio"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:U1")
    ws["A2"] = (
        "Edit yellow/blue score cells only (F–J). TOTAL = SUM(F:J). "
        "Priority = NOW/NEXT/LATER/PARK by TOTAL thresholds. "
        "Visibility: OPEN-CORE = public repo + paid layer on top; PRIVATE-PAID = closed source. "
        "Sorted recommendation: filter Priority = NOW."
    )
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    ws.merge_cells("A2:U2")

    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    # Data validation 1-5 for scores
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=False)
    dv.error = "Score must be integer 1–5"
    dv.errorTitle = "Invalid score"
    ws.add_data_validation(dv)

    for i, o in enumerate(OPPS):
        row = header_row + 1 + i
        values_pre = [
            o["id"], o["name"], o["category"], o["one_liner"], o["problem"],
        ]
        for col, v in enumerate(values_pre, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BLACK
            cell.border = THIN
            cell.alignment = WRAP if col in (4, 5) else LEFT
            if i % 2 == 1 and col not in range(6, 11):
                cell.fill = FILL_ALT

        # Scores F-J (cols 6-10) — blue + yellow
        scores = [o["burn"], o["soft_native"], o["ship"], o["dist"], o["moat"]]
        for j, s in enumerate(scores):
            col = 6 + j
            cell = ws.cell(row=row, column=col, value=s)
            cell.font = BLUE_INPUT
            cell.fill = FILL_YELLOW
            cell.alignment = CENTER
            cell.border = THIN
            dv.add(cell)

        # TOTAL formula col K
        total_cell = ws.cell(row=row, column=11, value=f"=F{row}+G{row}+H{row}+I{row}+J{row}")
        total_cell.font = BLACK_BOLD
        total_cell.alignment = CENTER
        total_cell.border = THIN

        # Priority formula col L
        # NOW if >=20, NEXT if >=16, LATER if >=12, else PARK
        pri_formula = (
            f'=IF(K{row}>="","",IF(K{row}>=20,"NOW",IF(K{row}>=16,"NEXT",IF(K{row}>=12,"LATER","PARK"))))'
        )
        # Actually K is always numeric; simplify
        pri_formula = f'=IF(K{row}>=20,"NOW",IF(K{row}>=16,"NEXT",IF(K{row}>=12,"LATER","PARK")))'
        pri_cell = ws.cell(row=row, column=12, value=pri_formula)
        pri_cell.font = BLACK_BOLD
        pri_cell.alignment = CENTER
        pri_cell.border = THIN

        vis, paid_layer = visibility_of(o)
        post = [
            vis, paid_layer,
            o["status"], o["stage"], o["platforms"], o["repo"], o["tags"], o["notes"], TODAY,
        ]
        for j, v in enumerate(post):
            col = 13 + j
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BLACK_BOLD if col == 13 else BLACK
            cell.border = THIN
            cell.alignment = WRAP if col in (14, 20) else LEFT
            if i % 2 == 1:
                cell.fill = FILL_ALT

        ws.row_dimensions[row].height = 48

    # Conditional formatting on Priority column
    last = header_row + len(OPPS)
    pri_range = f"L5:L{last}"
    ws.conditional_formatting.add(pri_range, FormulaRule(formula=['$L5="NOW"'], fill=FILL_NOW))
    ws.conditional_formatting.add(pri_range, FormulaRule(formula=['$L5="NEXT"'], fill=FILL_NEXT))
    ws.conditional_formatting.add(pri_range, FormulaRule(formula=['$L5="LATER"'], fill=FILL_LATER))
    ws.conditional_formatting.add(pri_range, FormulaRule(formula=['$L5="PARK"'], fill=FILL_PARK))

    # Visibility marker highlight: OPEN-CORE green-ish, PRIVATE-PAID grey
    vis_range = f"M5:M{last}"
    ws.conditional_formatting.add(vis_range, FormulaRule(formula=['$M5="OPEN-CORE"'], fill=FILL_NOW))
    ws.conditional_formatting.add(vis_range, FormulaRule(formula=['$M5="PRIVATE-PAID"'], fill=FILL_PARK))

    # Legend
    leg = last + 2
    ws.cell(row=leg, column=1, value="Legend").font = SUBTITLE_FONT
    ws.cell(row=leg + 1, column=1, value="Input scores").font = BLUE_INPUT
    ws.cell(row=leg + 1, column=1).fill = FILL_YELLOW
    ws.cell(row=leg + 1, column=2, value="Yellow fill + blue font = editable (1–5)").font = BLACK
    ws.cell(row=leg + 2, column=1, value="TOTAL / Priority").font = BLACK_BOLD
    ws.cell(row=leg + 2, column=2, value="Excel formulas — do not overwrite").font = BLACK
    ws.cell(row=leg + 3, column=1, value="NOW").fill = FILL_NOW
    ws.cell(row=leg + 3, column=2, value="TOTAL ≥ 20").font = BLACK
    ws.cell(row=leg + 4, column=1, value="NEXT").fill = FILL_NEXT
    ws.cell(row=leg + 4, column=2, value="TOTAL 16–19").font = BLACK
    ws.cell(row=leg + 5, column=1, value="LATER").fill = FILL_LATER
    ws.cell(row=leg + 5, column=2, value="TOTAL 12–15").font = BLACK
    ws.cell(row=leg + 6, column=1, value="PARK").fill = FILL_PARK
    ws.cell(row=leg + 6, column=2, value="TOTAL < 12").font = BLACK
    ws.cell(row=leg + 7, column=1, value="OPEN-CORE").fill = FILL_NOW
    ws.cell(row=leg + 7, column=2, value="Goes public (OSS repo); monetize via paid layer on top").font = BLACK
    ws.cell(row=leg + 8, column=1, value="PRIVATE-PAID").fill = FILL_PARK
    ws.cell(row=leg + 8, column=2, value="Stays private (closed source); the product itself is paid").font = BLACK

    # Freeze
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:U{last}"

    set_col_widths(ws, {
        "A": 6, "B": 22, "C": 22, "D": 48, "E": 42,
        "F": 8, "G": 14, "H": 11, "I": 12, "J": 8,
        "K": 9, "L": 10, "M": 14, "N": 30, "O": 14, "P": 14, "Q": 18, "R": 20, "S": 28, "T": 36, "U": 14,
    })


def build_deep_research(wb: Workbook) -> None:
    ws = wb.create_sheet("02_Deep_Research", 2)
    headers = [
        "ID", "Name", "Category", "Market context", "Competitors / alternatives",
        "Why now", "Risks", "Evidence / traction signals", "Next research questions",
        "Burn", "Software-native", "Ship", "Dist", "Moat", "TOTAL (link)",
    ]
    ws["A1"] = "Deep Research Briefs"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:O1")
    ws["A2"] = "Detailed research notes per opportunity. Score columns link to Master where noted; narrative is curated."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    for i, o in enumerate(OPPS):
        row = hr + 1 + i
        master_row = 5 + i
        data = [
            o["id"], o["name"], o["category"], o["market"], o["competitors"],
            o["why_now"], o["risks"], o["evidence"], o["next_research"],
            o["burn"], o["soft_native"], o["ship"], o["dist"], o["moat"],
        ]
        for col, v in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BLUE_INPUT if col >= 10 else BLACK
            cell.fill = FILL_YELLOW if col >= 10 else (FILL_ALT if i % 2 else FILL_WHITE)
            cell.border = THIN
            cell.alignment = WRAP if col in (4, 5, 6, 7, 8, 9) else (CENTER if col >= 10 else LEFT)
        # TOTAL link to master
        cell = ws.cell(row=row, column=15, value=f"='01_Master_Portfolio'!K{master_row}")
        cell.font = BLACK_BOLD
        cell.border = THIN
        cell.alignment = CENTER
        ws.row_dimensions[row].height = 72

    last = hr + len(OPPS)
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:O{last}"
    set_col_widths(ws, {
        "A": 6, "B": 22, "C": 20, "D": 40, "E": 36, "F": 32, "G": 32, "H": 32, "I": 32,
        "J": 8, "K": 12, "L": 8, "M": 8, "N": 8, "O": 12,
    })


def build_delivery(wb: Workbook) -> None:
    ws = wb.create_sheet("03_Build_Delivery", 3)
    headers = [
        "ID", "Name", "CLI", "Web", "Android", "iOS", "Desktop", "API", "Extension",
        "MVP scope", "Suggested stack", "Repo name", "Est. weeks", "Delivery notes", "Platforms (summary)",
    ]
    ws["A1"] = "Build & Delivery Matrix"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:O1")
    ws["A2"] = "Platform flags derived from opportunity design. MVP and stack are starting recommendations."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    platform_keys = ["CLI", "Web", "Android", "iOS", "Desktop", "API", "Extension"]

    for i, o in enumerate(OPPS):
        row = hr + 1 + i
        plats = {p.strip() for p in o["platforms"].split(",")}
        ws.cell(row=row, column=1, value=o["id"]).font = BLACK
        ws.cell(row=row, column=2, value=o["name"]).font = BLACK
        for j, pk in enumerate(platform_keys):
            col = 3 + j
            yes = pk in plats
            cell = ws.cell(row=row, column=col, value="Y" if yes else "—")
            cell.font = BLUE_INPUT if yes else BLACK
            cell.fill = FILL_YELLOW if yes else FILL_WHITE
            cell.alignment = CENTER
            cell.border = THIN
        for col, v in [
            (10, o["mvp"]),
            (11, o["stack"]),
            (12, o["repo"]),
            (13, o["effort_weeks"]),
            (14, o["delivery_notes"]),
            (15, o["platforms"]),
        ]:
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BLACK
            cell.border = THIN
            cell.alignment = WRAP if col in (10, 11, 14) else LEFT
            if col == 13:
                cell.font = BLUE_INPUT
                cell.fill = FILL_YELLOW
                cell.alignment = CENTER
        for c in range(1, 16):
            if ws.cell(row=row, column=c).border.left.style is None:
                ws.cell(row=row, column=c).border = THIN
            if i % 2 and c not in range(3, 10) and c != 13:
                if ws.cell(row=row, column=c).fill.fgColor is None or ws.cell(row=row, column=c).fill.fgColor.rgb in (
                    "00000000",
                    "0",
                ):
                    pass
        ws.row_dimensions[row].height = 56

    last = hr + len(OPPS)
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:O{last}"
    set_col_widths(ws, {
        "A": 6, "B": 22, "C": 6, "D": 6, "E": 9, "F": 6, "G": 9, "H": 6, "I": 10,
        "J": 42, "K": 32, "L": 20, "M": 10, "N": 32, "O": 18,
    })


def build_monetization(wb: Workbook) -> None:
    ws = wb.create_sheet("04_Monetization", 4)
    headers = [
        "ID", "Name", "Category", "ICP", "Pricing hypothesis", "GTM motion",
        "Revenue model", "ARPU hint", "Priority (link)", "TOTAL (link)",
    ]
    ws["A1"] = "Monetization & GTM"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    ws["A2"] = "ICP, pricing, and go-to-market per opportunity. Priority/TOTAL linked to Master."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    for i, o in enumerate(OPPS):
        row = hr + 1 + i
        master_row = 5 + i
        vals = [
            o["id"], o["name"], o["category"], o["icp"], o["pricing"],
            o["gtm"], o["revenue_model"], o["arpu_hint"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BLACK
            cell.border = THIN
            cell.alignment = WRAP if col in (4, 5, 6) else LEFT
            if i % 2:
                cell.fill = FILL_ALT
        # Priority & TOTAL links
        for col, letter in [(9, "L"), (10, "K")]:
            cell = ws.cell(row=row, column=col, value=f"='01_Master_Portfolio'!{letter}{master_row}")
            cell.font = BLACK_BOLD
            cell.border = THIN
            cell.alignment = CENTER
        # Pricing as editable input style on pricing hypothesis? Keep narrative black; ARPU blue
        arpu = ws.cell(row=row, column=8)
        arpu.font = BLUE_INPUT
        arpu.fill = FILL_YELLOW
        ws.row_dimensions[row].height = 56

    last = hr + len(OPPS)
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:J{last}"
    set_col_widths(ws, {
        "A": 6, "B": 22, "C": 22, "D": 36, "E": 36, "F": 36, "G": 16, "H": 16, "I": 12, "J": 12,
    })


def build_scoring_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("05_Scoring_Matrix", 5)
    headers = [
        "ID", "Name", "Burn", "Software-native", "Ship speed", "Distribution", "Moat", "TOTAL", "Priority",
    ]
    ws["A1"] = "Scoring Matrix (linked to Master)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = "All score and TOTAL cells are formulas → 01_Master_Portfolio. Change scores only on Master (or edit Master F–J)."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    for i, o in enumerate(OPPS):
        row = hr + 1 + i
        master_row = 5 + i
        ws.cell(row=row, column=1, value=o["id"]).font = BLACK
        ws.cell(row=row, column=2, value=o["name"]).font = BLACK
        # Link each score + total + priority
        for col, mcol in enumerate(["F", "G", "H", "I", "J", "K", "L"], 3):
            cell = ws.cell(row=row, column=col, value=f"='01_Master_Portfolio'!{mcol}{master_row}")
            cell.font = FORMULA_FONT
            cell.alignment = CENTER
            cell.border = THIN
        for c in range(1, 3):
            ws.cell(row=row, column=c).border = THIN
            ws.cell(row=row, column=c).alignment = LEFT

    last = hr + len(OPPS)

    # Color scale on TOTAL column H
    ws.conditional_formatting.add(
        f"H5:H{last}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )
    # Also color scale on score block C:G
    ws.conditional_formatting.add(
        f"C5:G{last}",
        ColorScaleRule(
            start_type="num", start_value=1, start_color="F8696B",
            mid_type="num", mid_value=3, mid_color="FFEB84",
            end_type="num", end_value=5, end_color="63BE7B",
        ),
    )
    ws.conditional_formatting.add(f"I5:I{last}", FormulaRule(formula=['$I5="NOW"'], fill=FILL_NOW))
    ws.conditional_formatting.add(f"I5:I{last}", FormulaRule(formula=['$I5="NEXT"'], fill=FILL_NEXT))
    ws.conditional_formatting.add(f"I5:I{last}", FormulaRule(formula=['$I5="LATER"'], fill=FILL_LATER))
    ws.conditional_formatting.add(f"I5:I{last}", FormulaRule(formula=['$I5="PARK"'], fill=FILL_PARK))

    # Bar chart of TOTAL by name — use data sorted? Chart from matrix as-is
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Opportunity TOTAL Score"
    chart.y_axis.title = "TOTAL (max 25)"
    chart.x_axis.title = "Opportunity"
    data = Reference(ws, min_col=8, min_row=4, max_row=last, max_col=8)
    cats = Reference(ws, min_col=2, min_row=5, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, "K4")

    # Rank helper table sorted note
    ws.cell(row=last + 2, column=1, value="Tip: Sort by TOTAL descending in Excel (Data → Sort) or filter Priority=NOW.").font = Font(
        name=FONT_NAME, size=9, italic=True
    )

    # Summary counts via COUNTIF on Priority
    ws.cell(row=last + 4, column=1, value="Pipeline counts").font = SUBTITLE_FONT
    for idx, band in enumerate(["NOW", "NEXT", "LATER", "PARK"]):
        ws.cell(row=last + 5 + idx, column=1, value=band).font = BLACK_BOLD
        cell = ws.cell(row=last + 5 + idx, column=2, value=f'=COUNTIF(I5:I{last},"{band}")')
        cell.font = BLACK
        cell.border = THIN

    ws.freeze_panes = "C5"
    set_col_widths(ws, {
        "A": 6, "B": 24, "C": 10, "D": 14, "E": 12, "F": 12, "G": 8, "H": 10, "I": 10, "K": 3,
    })


def build_pipeline(wb: Workbook) -> None:
    ws = wb.create_sheet("06_Pipeline", 6)
    ws["A1"] = "Delivery Pipeline by Priority Band"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "Phases populated from Master scores at generation time (static snapshot of assignment). "
        "Re-run generator after major score changes, or sort Master by Priority."
    )
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    def total(o):
        return o["burn"] + o["soft_native"] + o["ship"] + o["dist"] + o["moat"]

    def band(o):
        t = total(o)
        if t >= 20:
            return "NOW"
        if t >= 16:
            return "NEXT"
        if t >= 12:
            return "LATER"
        return "PARK"

    phases = {
        "NOW": ("Ship / double-down (next 30–60 days)", FILL_NOW),
        "NEXT": ("Validate & stage (60–120 days)", FILL_NEXT),
        "LATER": ("Backlog after NOW/NEXT capacity", FILL_LATER),
        "PARK": ("Watchlist — revisit quarterly", FILL_PARK),
    }

    headers = ["ID", "Name", "TOTAL", "Category", "Stage", "Repo", "Primary action"]
    actions = {
        "P01": "Grow npm adoption; launch posts (Show HN, r/cursor); suite cross-links",
        "P02": "Published — tighten checks; funnel from instruct-sync README",
        "P03": "Published — consolidate naming (@bhaskarauthor/mcp-sync vs mcp-config-sync); adoption",
        "P04": "Fold secret checks into P29 agent-skill-scan scope",
        "P05": "Superseded for NOW by P30; revisit as broad platform LATER",
        "P29": "Build next (≈3 wks) from setup-doctor parsers; launch on ToxicSkills narrative",
        "P30": "Design budget/kill-switch UX; 5 eng-manager interviews; build after P29",
        "P31": "Validate Hinglish ASR quality + ₹ WTP; needs cofounder/dedicated cycle",
        "P32": "Watch quarterly; build only if P29/P30 traction frees capacity",
        "P15": "Pick vertical; Meta BSP sandbox; 10 SMB interviews",
        "P17": "Salon/tutor interviews; WA reminder prototype",
    }

    r = 4
    for phase_name in ["NOW", "NEXT", "LATER", "PARK"]:
        title, fill = phases[phase_name]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(row=r, column=1, value=f"{phase_name} — {title}")
        cell.font = SECTION_FONT
        cell.fill = FILL_SECTION
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = FILL_SECTION
        r += 1
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = FILL_HEADER
            cell.border = THIN
            cell.alignment = CENTER
        r += 1
        group = [o for o in OPPS if band(o) == phase_name]
        group.sort(key=total, reverse=True)
        if not group:
            ws.cell(row=r, column=1, value="(none)").font = Font(name=FONT_NAME, size=10, italic=True)
            r += 2
            continue
        for o in group:
            t = total(o)
            vals = [
                o["id"], o["name"], t, o["category"], o["stage"], o["repo"],
                actions.get(o["id"], o["delivery_notes"]),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = BLACK_BOLD if col == 3 else BLACK
                cell.fill = fill
                cell.border = THIN
                cell.alignment = WRAP if col == 7 else LEFT
            # TOTAL also as formula reference optional — use static snapshot + note
            ws.row_dimensions[r].height = 30
            r += 1
        r += 2

    set_col_widths(ws, {
        "A": 6, "B": 24, "C": 8, "D": 24, "E": 16, "F": 20, "G": 48,
    })


def build_market_context(wb: Workbook) -> None:
    ws = wb.create_sheet("07_Market_Context", 7)
    ws["A1"] = "Market Context (Directional)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        "Macro notes for portfolio strategy. Figures are directional mid-decade class estimates for brainstorming—"
        "not audited financial advice. Validate with primary sources before investor materials."
    )
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    ws.merge_cells("A2:F2")

    headers = ["Theme", "Directional figure", "Implication for portfolio", "Related IDs", "Source / note class", "Confidence"]
    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    rows = [
        (
            "AI agents / agent platforms",
            "~$7–11B market mid-decade class; high CAGR ~45% class in bullish analyst ranges",
            "Prioritize AI-native tooling with PLG (CLI/npx) while category is forming",
            "P01–P08",
            "Industry analyst ranges (e.g. agent/AI software market notes 2024–26); treat as order-of-magnitude",
            "Medium",
        ),
        (
            "India digital economy / DX",
            "India DX ~$124–144B class for 2025–26 windows in public industry estimates; SME high CAGR",
            "India SMB WhatsApp/GST/appointments/clinic track is strategic, not side quest",
            "P15–P18, P16",
            "Industry/NASSCOM-class digital economy commentary; range compressed for planning",
            "Medium",
        ),
        (
            "Global SaaS",
            "SaaS still hundreds of $B TAM class; SME adoption continues despite efficiency era",
            "Seat/MRR models remain valid; need sharp ICP and low CAC",
            "P05, P11, P14, P21",
            "Public SaaS market sizing consensus (broad)",
            "Medium",
        ),
        (
            "Cloud waste / FinOps",
            "~30% cloud spend waste class repeatedly cited in FinOps/Flexera-style surveys",
            "Cloud waste radar viable but crowded—need niche or automation wedge",
            "P11, P05",
            "Flexera State of the Cloud-class surveys (historical ~30% waste narrative)",
            "Medium",
        ),
        (
            "AI cybersecurity / managed services",
            "Large and growing; AI expands both attack surface and defense tooling",
            "Shadow-AI, secret-guard, cyber-SMB are real but sales/liability heavy",
            "P04, P07, P20, P09",
            "Cybersecurity market reports (AI security subsegment growing faster than overall)",
            "Medium",
        ),
        (
            "MCP & multi-agent coding",
            "Rapid ecosystem growth; standard still evolving (2024–2026)",
            "instruct-sync / mcp-sync / setup-doctor ride open standards + multi-vendor pain",
            "P01–P04, P06",
            "Primary: MCP adoption in Cursor/VS Code/Claude ecosystems; GitHub discussions",
            "High (qualitative)",
        ),
        (
            "Open source distribution",
            "npx/npm cold-start remains elite PLG for developer tools",
            "Bias portfolio toward OSS CLI NOW products with optional cloud later",
            "P01, P02, P03, P04, P08, P13, P28",
            "Empirical: successful DX CLIs (eslint, prettier, supabase CLI patterns)",
            "High",
        ),
        (
            "India WhatsApp commerce / ops",
            "WhatsApp is default business OS for huge SMB base; Cloud API monetizable",
            "wa-ops-desk + appt-book are GTM-native for Bharat SMB",
            "P15, P17, P18",
            "Meta WhatsApp Business / India SMB digitization narratives",
            "Medium-High",
        ),
        (
            "Compliance / GRC automation",
            "SOC2/ISO tooling multi-player market; India DPDP creates local angle",
            "Avoid pure Vanta clone; consider DPDP or AI-evidence niche only",
            "P14, P07",
            "Category defined by Vanta/Drata/Sprinto success",
            "High (category), Low (our wedge)",
        ),
        (
            "Creator & consumer apps",
            "Large attention markets; brutal retention and CAC",
            "Keep focus-forge, care-companion, creator-ops as PARK unless passion leverage",
            "P22–P24, P27",
            "Consumer app economics (general)",
            "High",
        ),
        (
            "Agent supply-chain security",
            "36.8% of ~4k marketplace skills flawed, 1,467 malicious payloads (Snyk ToxicSkills); 71% of MCP servers graded F; 29% of orgs prepared",
            "agent-skill-scan (P29) is the highest-conviction new NOW; absorbs secret-guard scope",
            "P29, P04, P02",
            "Snyk ToxicSkills study Feb 2026; MCP server audit 2026; incident reporting",
            "High",
        ),
        (
            "AI token cost crisis",
            "Teams 3x over annual token budgets by April 2026 (FinOps Foundation); Uber blew 2026 AI coding budget by April; 85% of GenAI deployments unobserved",
            "Guardrails/enforcement wedge (P30) beats another dashboard; per-seat prosumer pricing viable",
            "P30, P05",
            "TechCrunch Jun 2026; FinOps Foundation commentary; observability market reports",
            "High",
        ),
        (
            "Paid consumer voice AI",
            "Wispr Flow ~$10M ARR, $700M valuation, $315M raised — English-first",
            "Hinglish/Indic dictation (P31) is the underserved adjacent; don't fight Wispr in English",
            "P31",
            "Public funding/revenue reporting 2025–26",
            "Medium-High",
        ),
    ]

    for i, row_data in enumerate(rows):
        row = hr + 1 + i
        for col, v in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BLACK
            cell.border = THIN
            cell.alignment = WRAP
            if i % 2:
                cell.fill = FILL_ALT
        ws.row_dimensions[row].height = 48

    note_row = hr + len(rows) + 2
    ws.cell(row=note_row, column=1, value="Disclaimer").font = SUBTITLE_FONT
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 2, end_column=6)
    ws.cell(
        row=note_row + 1,
        column=1,
        value=(
            "Numbers are planning ranges for brainstorming alignment, not forecasts. "
            "Re-source from Gartner/IDC/NASSCOM/Flexera/primary filings before external use. "
            "Last portfolio refresh: " + TODAY + "."
        ),
    ).font = Font(name=FONT_NAME, size=9, italic=True)
    ws.cell(row=note_row + 1, column=1).alignment = WRAP

    set_col_widths(ws, {"A": 28, "B": 42, "C": 42, "D": 18, "E": 42, "F": 14})


def build_research_log(wb: Workbook) -> None:
    ws = wb.create_sheet("08_Research_Log", 8)
    ws["A1"] = "Research Log"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws["A2"] = "Append-only style log. Yellow/blue = editable fields for new entries."
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color="666666")

    headers = ["Date", "Opportunity ID", "Type", "Summary", "Source", "Impact on scores?", "Follow-up"]
    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN

    starter = [
        (
            "2026-07-09",
            "P01",
            "Product",
            "instruct-sync public beta on npm as ai-instruct-sync; multi-agent rules sync working",
            "Local repo / npm beta",
            "Confirmed high ship+dist scores",
            "Show HN + retention metrics",
        ),
        (
            "2026-07-09",
            "P02",
            "Product",
            "ai-setup-doctor scaffolded build-green with fixtures for MCP/secrets/drift",
            "Local repo ai-setup-doctor/",
            "Supports NOW priority",
            "npm publish + README funnel from instruct-sync",
        ),
        (
            "2026-07-09",
            "PORTFOLIO",
            "Meta",
            "Initial 28-opportunity portfolio scored; AI CLI cluster prioritized",
            "Brainstorm + workspace roadmap",
            "Baseline scores set",
            "Re-score after 5 user interviews per NOW item",
        ),
        (
            "2026-07-09",
            "P03",
            "Market",
            "MCP multi-client config drift remains open problem; sister tool to instruct-sync justified",
            "Ecosystem observation / roadmap",
            "NEXT candidate",
            "Enumerate client MCP paths",
        ),
        (
            "2026-07-09",
            "P15",
            "Market",
            "India SMB WhatsApp ops remains high burn; competitive BSP landscape noted",
            "Market context notes",
            "Burn high, moat low",
            "Vertical selection interviews",
        ),
        (
            "2026-07-09",
            "P11",
            "Market",
            "Cloud waste ~30% class from FinOps surveys—category real but crowded",
            "Flexera-class survey narrative",
            "Moat scored low",
            "Only pursue with niche wedge",
        ),
        (
            "2026-07-09",
            "MACRO",
            "Market",
            "AI agents ~$7–11B mid-decade / ~45% CAGR class directional; India DX ~$124–144B class",
            "Industry estimate ranges (directional)",
            "Supports AI tooling + India SMB dual track",
            "Attach primary sources when externalizing",
        ),
        (
            "2026-07-11",
            "P02",
            "Product",
            "Verified published on npm: ai-setup-doctor@0.1.0-beta.0 (latest+beta)",
            "npm registry check",
            "Status → Published beta; stage → Shipped",
            "Launch posts; funnel from instruct-sync README",
        ),
        (
            "2026-07-11",
            "P03",
            "Product",
            "Verified published on npm: @bhaskarauthor/mcp-sync@0.2.0 + alias mcp-config-sync@0.2.0",
            "npm registry check",
            "Status → Published; ship 4→5 (TOTAL 21)",
            "Consolidate package naming; adoption metrics",
        ),
        (
            "2026-07-11",
            "P29",
            "Market",
            "Agent skills supply chain burning: ToxicSkills 1,467 malicious payloads, 36.8% flawed; 71% MCP servers grade F — added agent-skill-scan (TOTAL 23, NOW)",
            "Snyk ToxicSkills; MCP audit 2026",
            "New NOW item; absorbs P04 scope",
            "Injection heuristic corpus test; build ≈3 wks",
        ),
        (
            "2026-07-11",
            "P30",
            "Market",
            "Token bill crisis (teams 3x over 2026 budgets by April; Uber, Microsoft anecdotes) — added agent-spend-guard budgets/kill-switch (TOTAL 21, NOW)",
            "TechCrunch Jun 2026; FinOps Foundation",
            "New NOW item; supersedes P05 for NOW",
            "Eng-manager WTP interviews; wrapper vs proxy",
        ),
        (
            "2026-07-11",
            "P31",
            "Market",
            "Wispr Flow ~$10M ARR/$700M valuation validates paid dictation; Hinglish/Indic wedge added as indic-voice-flow (TOTAL 16, NEXT, PRIVATE-PAID)",
            "Public funding/revenue reporting",
            "New NEXT consumer revenue bet",
            "Hinglish ASR quality test; 20 professional interviews",
        ),
        (
            "2026-07-11",
            "PORTFOLIO",
            "Meta",
            "Added Visibility marker (OPEN-CORE vs PRIVATE-PAID) + Paid layer column across all 32 items; hygiene-suite strategic line adopted (P01→P03→P02→P29→P30)",
            "Portfolio review session",
            "Structure change; no re-scores except P03 ship",
            "Re-score after launch-week signals",
        ),
        (
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ),
        (
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ),
        (
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ),
    ]

    for i, row_data in enumerate(starter):
        row = hr + 1 + i
        for col, v in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=v if v != "" else None)
            cell.font = BLUE_INPUT
            cell.fill = FILL_YELLOW
            cell.border = THIN
            cell.alignment = WRAP if col in (4, 7) else LEFT
        ws.row_dimensions[row].height = 36

    # Extra empty input rows
    for i in range(10):
        row = hr + 1 + len(starter) + i
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col, value=None)
            cell.font = BLUE_INPUT
            cell.fill = FILL_YELLOW
            cell.border = THIN

    set_col_widths(ws, {"A": 12, "B": 14, "C": 10, "D": 56, "E": 28, "F": 24, "G": 32})
    ws.freeze_panes = "A5"


def write_csv() -> None:
    fieldnames = [
        "id", "name", "category", "one_liner", "problem",
        "burn", "software_native", "ship_speed", "distribution", "moat",
        "total", "priority", "visibility", "paid_layer",
        "status", "stage", "platforms", "repo", "tags", "notes",
        "last_reviewed",
    ]

    def band_total(o):
        t = o["burn"] + o["soft_native"] + o["ship"] + o["dist"] + o["moat"]
        if t >= 20:
            p = "NOW"
        elif t >= 16:
            p = "NEXT"
        elif t >= 12:
            p = "LATER"
        else:
            p = "PARK"
        return t, p

    with CSV_PATH.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for o in OPPS:
            t, p = band_total(o)
            vis, paid_layer = visibility_of(o)
            w.writerow({
                "id": o["id"],
                "name": o["name"],
                "category": o["category"],
                "one_liner": o["one_liner"],
                "problem": o["problem"],
                "burn": o["burn"],
                "software_native": o["soft_native"],
                "ship_speed": o["ship"],
                "distribution": o["dist"],
                "moat": o["moat"],
                "total": t,
                "priority": p,
                "visibility": vis,
                "paid_layer": paid_layer,
                "status": o["status"],
                "stage": o["stage"],
                "platforms": o["platforms"],
                "repo": o["repo"],
                "tags": o["tags"],
                "notes": o["notes"],
                "last_reviewed": TODAY,
            })


def write_index() -> None:
    # compute top NOW
    scored = []
    for o in OPPS:
        t = o["burn"] + o["soft_native"] + o["ship"] + o["dist"] + o["moat"]
        if t >= 20:
            band = "NOW"
        elif t >= 16:
            band = "NEXT"
        elif t >= 12:
            band = "LATER"
        else:
            band = "PARK"
        scored.append((t, band, o))
    scored.sort(key=lambda x: (-x[0], x[2]["id"]))
    now = [s for s in scored if s[1] == "NOW"]

    lines = [
        "# Software Opportunity Portfolio — Index",
        "",
        f"**Last generated:** {TODAY}",
        "",
        "## Files",
        "",
        f"| File | Description |",
        f"|------|-------------|",
        f"| [`Software_Opportunity_Research_Portfolio.xlsx`](./Software_Opportunity_Research_Portfolio.xlsx) | Full multi-sheet research workbook (openpyxl) |",
        f"| [`Software_Opportunity_Master_Portfolio.csv`](./Software_Opportunity_Master_Portfolio.csv) | Flat export of all {len(OPPS)} opportunities with scores + visibility markers |",
        f"| [`SOFTWARE_PORTFOLIO_INDEX.md`](./SOFTWARE_PORTFOLIO_INDEX.md) | This index |",
        f"| [`scripts/generate_software_portfolio.py`](./scripts/generate_software_portfolio.py) | Regenerator script |",
        "",
        "## Workbook sheets",
        "",
        "1. **00_README** — how to use, scoring legend, conventions",
        "2. **01_Master_Portfolio** — all opportunities; editable scores (yellow/blue); `TOTAL` and `Priority` formulas",
        "3. **02_Deep_Research** — market, competitors, risks, evidence",
        "4. **03_Build_Delivery** — platforms (CLI/Web/Android/iOS/Desktop/API/Extension), MVP, stack, repos",
        "5. **04_Monetization** — ICP, pricing, GTM",
        "6. **05_Scoring_Matrix** — cross-sheet formulas, color scale, bar chart",
        "7. **06_Pipeline** — NOW / NEXT / LATER / PARK phases",
        "8. **07_Market_Context** — macro notes with source classes",
        "9. **08_Research_Log** — research log with starter entries",
        "",
        "## Scoring",
        "",
        "Dimensions (1–5): **Burn**, **Software-native**, **Ship speed**, **Distribution**, **Moat**.",
        "",
        "- **TOTAL** = sum of five scores (Excel: `=F+G+H+I+J`)",
        "- **Priority:** NOW ≥ 20 · NEXT 16–19 · LATER 12–15 · PARK < 12",
        "- Inputs: **blue font + yellow fill** (financial modeling convention)",
        "",
        "## Visibility markers",
        "",
        "- **OPEN-CORE** — goes public (OSS repo, npx distribution); money made on a paid layer on top (CI action, team/policy tier, hosted service)",
        "- **PRIVATE-PAID** — stays private (closed source); the product itself is the paid thing",
        "",
        "## Strategic line — AI dev hygiene suite",
        "",
        "One audience (AI-assisted developers), one flywheel (npx cold-start), five OPEN-CORE products that cross-market each other; paid layers on top:",
        "",
        "1. **Sync** — P01 instruct-sync (shipped) + P03 mcp-sync (shipped)",
        "2. **Diagnose** — P02 ai-setup-doctor (shipped)",
        "3. **Secure** — P29 agent-skill-scan (build next, ≈3 wks)",
        "4. **Control spend** — P30 agent-spend-guard (design; the suite's money product)",
        "",
        "Standalone revenue bet outside the suite: P31 indic-voice-flow (PRIVATE-PAID, NEXT).",
        "",
        "## Top NOW priorities (by TOTAL)",
        "",
    ]
    if now:
        lines.append("| Rank | ID | Name | TOTAL | Visibility | Notes |")
        lines.append("|------|----|------|-------|------------|-------|")
        for i, (t, _, o) in enumerate(now, 1):
            vis, _paid = visibility_of(o)
            lines.append(f"| {i} | {o['id']} | {o['name']} | {t} | {vis} | {o['notes']} |")
    else:
        lines.append("_No NOW items at current thresholds._")

    lines += [
        "",
        "## Regenerate",
        "",
        "```bash",
        "cd /Users/bhaskar_pandey/Documents/development",
        "python3 scripts/generate_software_portfolio.py",
        "```",
        "",
        "## Related workspace projects",
        "",
        "- `instruct-sync/` — P01 (npm `ai-instruct-sync@0.2.0-beta.0`)",
        "- `ai-setup-doctor/` — P02 (npm `ai-setup-doctor@0.1.0-beta.0`)",
        "- `mcp-sync/` — P03 (npm `mcp-config-sync@0.2.0` + `@bhaskarauthor/mcp-sync`; own git remote, co-located)",
        "- `ai-agent-roadmap.md` — 90-day AI tooling plan",
        "",
    ]
    INDEX_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    build_readme(wb)
    build_master(wb)
    build_deep_research(wb)
    build_delivery(wb)
    build_monetization(wb)
    build_scoring_matrix(wb)
    build_pipeline(wb)
    build_market_context(wb)
    build_research_log(wb)

    wb.save(XLSX_PATH)
    write_csv()
    write_index()
    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote {CSV_PATH}")
    print(f"Wrote {INDEX_PATH}")

    # Print NOW priorities
    print("\nNOW priorities:")
    rows = []
    for o in OPPS:
        t = o["burn"] + o["soft_native"] + o["ship"] + o["dist"] + o["moat"]
        if t >= 20:
            rows.append((t, o["id"], o["name"]))
    for t, i, n in sorted(rows, reverse=True):
        print(f"  {i} {n}: {t}")


if __name__ == "__main__":
    main()
